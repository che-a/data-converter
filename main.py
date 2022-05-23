#!/usr/bin/env python3
"""
Скрипт преобразования данных из файла Excel в файл JSON.
"""
import argparse
import json

from openpyxl import load_workbook

DEFAULT_SETTINGS_FILE = "settings.json"
DEFAULT_INPUT_FILE = "input.xlsx"
DEFAULT_OUTPUT_FILE = "output.json"

ERROR_NOT_FOUND_SETTINGS_FILE = 2
ERROR_NOT_FOUND_INPUT_FILE = 3


def load_settings_file(filename=DEFAULT_SETTINGS_FILE):
    """
    Загрузка JSON-файла с настройками
    :param filename: имя JSON-файла
    :return: загруженный из JSON-файла словарь
    """
    with open(filename) as json_file:
        result = json.load(json_file)
    return result


def load_excel_file(limits_dict, filename):
    """
    Загрузка содержимого Excel-файла согласно настроек ограничений.
    :param limits_dict: словарь с настройками ограничений.
    :param filename: имя входящего Excel-файла.
    :return: словарь с данными.
    """
    wb = load_workbook(filename)
    ws = wb[limits_dict['sheet']]

    result = {}
    for row in range(limits_dict['begin_row'], limits_dict['end_row'] + 1):
        if row not in limits_dict['exclude_rows']:
            unique_key = str(ws[limits_dict['unique_key_column'] + str(row)].value)
            result[unique_key] = {}

            for column_name, column_value in limits_dict['columns'].items():
                if column_value != limits_dict['unique_key_column']:
                    result[unique_key][column_name] = ws[column_value + str(row)].value
    return result


def save_json_file(data_dict, filename):
    with open(filename, 'w') as outfile:
        json.dump(data_dict, outfile, indent=4, ensure_ascii=False)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("-i", dest="input_file", default=DEFAULT_INPUT_FILE, type=str, help="Входной Excel-файл")
    parser.add_argument("-o", dest="output_file", default=DEFAULT_OUTPUT_FILE, type=str, help="Выходной JSON-файл")
    args = parser.parse_args()

    # Чтение файла с настройками
    limits = dict()
    try:
        limits = load_settings_file()
    except FileNotFoundError:
        print(f"ОШИБКА: не найден файл настроек {DEFAULT_SETTINGS_FILE}")
        exit(ERROR_NOT_FOUND_SETTINGS_FILE)

    # Чтение входного Excel-файла
    data = dict()
    try:
        data = load_excel_file(limits, args.input_file)
    except FileNotFoundError:
        print(f"ОШИБКА: не найден файл {args.input_file}")
        exit(ERROR_NOT_FOUND_INPUT_FILE)

    # Сохранение JSON-файла с выгруженными данными
    save_json_file(data, args.output_file)
